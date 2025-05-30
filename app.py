import io
import os
import traceback
from datetime import datetime
from flask import (
    Flask,
    request,
    render_template,
    send_file,
    url_for
)
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import copy
import re # 正規表現のためにreモジュールをインポート

# PDF処理用ライブラリ
try:
    import pdfplumber  # 表構造を保持したPDF処理に最適
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False
    # フォールバック用にPyPDF2も保持
    try:
        from PyPDF2 import PdfReader
    except ImportError:
        from PyPDF2 import PdfFileReader

app = Flask(__name__, template_folder='templates', static_folder='static')

# --- 定数設定 ---
TEMPLATE_FILE_PATH = 'template.xlsm'  # テンプレートファイルのパス
ALLOWED_EXTENSIONS = {'xls', 'xlsx', 'xlsm', 'pdf'} # xlsmも追加
MAX_CONTENT_LENGTH = 32 * 1024 * 1024  # 32MBに増量（PDFサイズ考慮）

# --- 初期設定 ---
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

def allowed_file(filename: str) -> bool:
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def extract_pdf_tables(pdf_bytes, filename):
    """
    PDFから表構造を保持してデータを抽出する関数
    pdfplumberを使用して表の縦線・横線を認識し、セル構造を維持する
    """
    extracted_data = []
    
    if PDFPLUMBER_AVAILABLE:
        try:
            # pdfplumberを使用した高度な表抽出
            with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    page_data = {
                        'page_number': page_num + 1,
                        'tables': [],
                        'text': ''
                    }
                    
                    # ページから表を抽出
                    tables = page.extract_tables()
                    
                    if tables:
                        # 表が見つかった場合
                        for table_idx, table in enumerate(tables):
                            if table and len(table) > 0:
                                # 空の行やセルを適切に処理
                                clean_table = []
                                for row in table:
                                    if row:  # 行が空でない場合
                                        clean_row = []
                                        for cell in row:
                                            # セルの内容を文字列として処理
                                            if cell is not None:
                                                # 改行や余分な空白を整理
                                                cell_content = str(cell).strip().replace('\n', ' ')
                                                clean_row.append(cell_content)
                                            else:
                                                clean_row.append('')  # 空のセル
                                        clean_table.append(clean_row)
                                
                                if clean_table:  # 空でない表のみ保存
                                    page_data['tables'].append({
                                        'table_index': table_idx,
                                        'data': clean_table
                                    })
                    
                    # 表以外のテキストも抽出（補足情報として）
                    page_text = page.extract_text()
                    if page_text:
                        # 改行を適切に処理
                        clean_text = page_text.strip().replace('\r\n', '\n').replace('\r', '\n')
                        page_data['text'] = clean_text
                    
                    extracted_data.append(page_data)
            
        except Exception as e:
            print(f"pdfplumberでの処理中にエラーが発生: {e}")
            # フォールバック処理
            return extract_pdf_fallback(pdf_bytes, filename)
    
    else:
        # pdfplumberが利用できない場合のフォールバック
        return extract_pdf_fallback(pdf_bytes, filename)
    
    return extracted_data

def extract_pdf_fallback(pdf_bytes, filename):
    """
    pdfplumberが利用できない場合のフォールバック処理
    PyPDF2を使用した基本的なテキスト抽出
    """
    extracted_data = []
    
    try:
        pdf_reader = PdfReader(io.BytesIO(pdf_bytes))
        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            text = page.extract_text()
            
            page_data = {
                'page_number': page_num + 1,
                'tables': [],
                'text': text.strip() if text else f"[ページ {page_num + 1}: テキストを抽出できませんでした]"
            }
            extracted_data.append(page_data)
    
    except Exception as e:
        print(f"PyPDF2での処理中にエラーが発生: {e}")
        # 最後の手段として空のページデータを返す
        extracted_data.append({
            'page_number': 1,
            'tables': [],
            'text': f"[エラー: {filename} を処理できませんでした - {str(e)}]"
        })
    
    return extracted_data

def copy_worksheet_data(source_sheet, target_sheet, start_row, start_col):
    """
    ソースシートからターゲットシートにデータをコピーする関数
    既存のデータがある場合は指定した位置から貼り付ける
    """
    # ソースシートの使用範囲を取得
    if source_sheet.max_row == 1 and source_sheet.max_column == 1:
        # 空のシートの場合はスキップ
        return
    
    # データをコピー（値のみ、書式は保持しない）
    for row in range(1, source_sheet.max_row + 1):
        for col in range(1, source_sheet.max_column + 1):
            source_cell = source_sheet.cell(row=row, column=col)
            if source_cell.value is not None:  # 値がある場合のみコピー
                target_cell = target_sheet.cell(
                    row=start_row + row - 1, 
                    column=start_col + col - 1
                )
                target_cell.value = source_cell.value

# --- ルーティング ---

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload_and_process', methods=['POST'])
def upload_and_process():
    try:
        # 1. ファイルの存在確認
        if 'excel_file' not in request.files or 'pdf_files' not in request.files:
            return render_template('error.html', message="ExcelファイルとPDFファイルの両方をアップロードしてください。"), 400

        excel_file_storage = request.files['excel_file']
        pdf_files_storage = request.files.getlist('pdf_files') # 複数PDFを受け取る

        if not excel_file_storage.filename:
            return render_template('error.html', message="Excelファイルが選択されていません。"), 400
        if not all(f.filename for f in pdf_files_storage):
            return render_template('error.html', message="PDFファイルが選択されていません。"), 400
        
        # 2. テンプレートファイルの存在確認と読み込み
        if not os.path.exists(TEMPLATE_FILE_PATH):
            return render_template('error.html', message=f"テンプレートファイル '{TEMPLATE_FILE_PATH}' が見つかりません。"), 500
        
        # テンプレートファイルを読み込み（元のファイルは変更されない）
        template_workbook = load_workbook(TEMPLATE_FILE_PATH, keep_vba=True)
        
        # 3. アップロードされたExcelファイルの読み込み
        uploaded_excel_workbook = load_workbook(io.BytesIO(excel_file_storage.read()))
        
        # アップロードされたExcelの1ページ目を取得
        if len(uploaded_excel_workbook.worksheets) == 0:
            return render_template('error.html', message="アップロードされたExcelファイルにシートがありません。"), 400
        
        uploaded_first_sheet = uploaded_excel_workbook.worksheets[0]
        
        # 4. テンプレートの1枚目のシートにアップロードされたデータを貼り付け
        if len(template_workbook.worksheets) == 0:
            return render_template('error.html', message="テンプレートファイルにシートがありません。"), 500
        
        template_first_sheet = template_workbook.worksheets[0]
        
        # 既存データの最後の行を取得
        last_row = template_first_sheet.max_row if template_first_sheet.max_row > 1 else 1
        
        # 空の行があるかチェック
        is_empty_template = True
        for row in range(1, last_row + 1):
            for col in range(1, template_first_sheet.max_column + 1):
                if template_first_sheet.cell(row=row, column=col).value is not None:
                    is_empty_template = False
                    break
            if not is_empty_template:
                break
        
        # データを貼り付ける開始位置を決定
        if is_empty_template:
            start_row = 1  # テンプレートが空の場合は1行目から
        else:
            start_row = last_row + 2  # 既存データがある場合は2行空けて貼り付け
        
        copy_worksheet_data(uploaded_first_sheet, template_first_sheet, start_row, 1)

        # 5. PDFファイルの処理（表構造を保持する高度な処理）
        for pdf_file_storage in pdf_files_storage:
            pdf_bytes = pdf_file_storage.read()
            
            # 高度なPDF表抽出を実行
            pdf_data = extract_pdf_tables(pdf_bytes, pdf_file_storage.filename)
            
            if not pdf_data:
                return render_template('error.html', 
                    message=f"PDF '{pdf_file_storage.filename}' から有効な内容を抽出できませんでした。"), 500

            # 抽出したデータをテンプレートの新しいシートに配置
            for page_data in pdf_data:
                page_num = page_data['page_number']
                
                # シート名を「Page_1」「Page_2」の形式で生成
                sheet_name = f"Page_{page_num}"
                
                # 同名のシートが既に存在する場合は番号を付ける
                counter = 1
                while sheet_name in [ws.title for ws in template_workbook.worksheets]:
                    sheet_name = f"Page_{page_num}_{counter}"
                    counter += 1

                # 新しいシートを作成
                new_sheet = template_workbook.create_sheet(title=sheet_name)
                
                # --- PDFテキストから特定情報を抽出 ---
                extracted_title = ""
                extracted_name = ""
                extracted_role = ""
                
                if page_data['text']:
                    lines = page_data['text'].split('\n')
                    # 各行が抽出に使用されたかどうかを追跡するフラグ
                    line_used_for_extraction = [False] * len(lines) 

                    for i, line in enumerate(lines):
                        current_line = line.strip()
                        if not current_line: # 空行はスキップし、後で「その他のテキスト」にも含めない
                            line_used_for_extraction[i] = True 
                            continue

                        # 1. タイトル（例: "YYYY年M月出勤簿"）を抽出
                        if not extracted_title: # まだタイトルが抽出されていない場合のみ試行
                            # 年、月、出勤簿の間にスペースを許容する正規表現
                            title_match = re.search(r'(\d{4}年\s*\d{1,2}月\s*出勤簿)', current_line)
                            if title_match:
                                extracted_title = title_match.group(1).strip()
                                line_used_for_extraction[i] = True
                                continue # タイトルが見つかったら、この行での他の抽出は行わない

                        # 2. 雇用と役割（例: "正社員（一般）・園長（配置内）"）を抽出
                        if not extracted_role: # まだ役割が抽出されていない場合のみ試行
                            role_keywords = ["正社員", "パート", "園長", "保育士"]
                            found_roles_on_line = []
                            for role_kw in role_keywords:
                                # 単語の境界(\b)を使って、より正確にキーワードをマッチさせ、全てのマッチを収集
                                matches = re.findall(r'\b' + re.escape(role_kw) + r'\b', current_line)
                                if matches:
                                    for match in matches:
                                        if match not in found_roles_on_line: # 重複防止
                                            found_roles_on_line.append(match)

                            if found_roles_on_line:
                                # 見つかった役割をソートし、ユニークなものだけを " / " で結合
                                extracted_role = " / ".join(sorted(list(set(found_roles_on_line))))
                                line_used_for_extraction[i] = True
                                continue # 役割が見つかったら、この行での他の抽出は行わない

                        # 3. 名前を抽出
                        # 優先度順に試行
                        if not extracted_name: # まだ名前が抽出されていない場合のみ試行
                            # 優先度1 (修正): 10桁のID番号に続く名前（例: "2025030040 日高 安澄"）
                            # ID番号の後に日本語の名前があり、その後には日付情報が続く可能性があることを考慮
                            match_id_name = re.search(
                                r'^\s*\d{10}\s*([\u3000-\u9FFF\uFF00-\uFFEF\s・ー]+?)(?:\s+\d{4}年|\s*$)', current_line
                            )
                            if match_id_name:
                                potential_name = match_id_name.group(1).strip()
                                # 抽出された名前が日付や役割/タイトルキーワードを含まないか最終確認
                                if not re.search(r'\d{4}年|\d{1,2}月|\d{1,2}日|正社員|パート|園長|保育士|出勤簿', potential_name):
                                    extracted_name = potential_name
                                    line_used_for_extraction[i] = True
                                    continue # 名前が見つかったら、この行での他の名前検索は不要

                            # 優先度2: "氏名："に続く名前 (既存ロジックをフォールバックとして残す)
                            match_shimei = re.search(r'氏名[:：\s]*([\u3000-\u9FFF\uFF00-\uFFEF\s・ー]{2,})', current_line)
                            if match_shimei:
                                potential_name = match_shimei.group(1).strip()
                                if not re.search(r'\d{4}年|\d{1,2}月|\d{1,2}日|正社員|パート|園長|保育士|出勤簿', potential_name):
                                    extracted_name = potential_name
                                    line_used_for_extraction[i] = True
                                    continue 

                            # 優先度3: 日付に挟まれた名前 (既存ロジックをフォールバックとして残す)
                            match_name_between_dates = re.search(
                                r'\d{4}年\d{1,2}月\d{1,2}日\s*([\u3000-\u9FFF\uFF00-\uFFEF\s・ー]+?)\s*\d{4}年\d{1,2}月\d{1,2}日', current_line
                            )
                            if match_name_between_dates:
                                potential_name = match_name_between_dates.group(1).strip()
                                if not re.search(r'\d{4}年|\d{1,2}月|\d{1,2}日|正社員|パート|園長|保育士|出勤簿', potential_name):
                                    extracted_name = potential_name
                                    line_used_for_extraction[i] = True
                                    continue 

                            # 優先度4: 単独の日本語名 (既存ロジックをフォールバックとして残す)
                            # 行全体が日付パターンや役割/タイトルキーワードを含まないことを確認
                            if not re.search(r'\d{4}年|\d{1,2}月|\d{1,2}日|正社員|パート|園長|保育士|出勤簿', current_line):
                                match_name_standalone = re.search(r'^([\u3000-\u9FFF\uFF00-\uFFEF\s・ー]{2,})$', current_line)
                                if match_name_standalone:
                                    extracted_name = match_name_standalone.group(1).strip()
                                    line_used_for_extraction[i] = True
                                    continue 

                    # 抽出に使用されなかった残りのテキスト行を収集
                    remaining_text_lines = []
                    for i, line in enumerate(lines):
                        if not line_used_for_extraction[i] and line.strip(): # 使用されていない、かつ空でない行のみ追加
                            remaining_text_lines.append(line.strip())

                # page_data['text']を、抽出されなかった残りのテキストのみで更新
                page_data['text'] = '\n'.join(remaining_text_lines) if remaining_text_lines else ''

                # --- 抽出した情報をExcelシートの特定セルに配置 ---
                # A1セルにタイトル
                new_sheet.cell(row=1, column=1, value=extracted_title) 
                # A2セルに名前
                new_sheet.cell(row=2, column=1, value=extracted_name)  
                # C2セルに役割
                new_sheet.cell(row=2, column=3, value=extracted_role)  

                # 以降のコンテンツ（表や残りのテキスト）の開始行を調整
                # A1, A2, C2のために2行使用したので、3行目以降から開始
                current_row = 4 # 3行目まで使用済みの場合は4行目から開始
                
                # 表データがある場合は表として配置
                if page_data['tables']:
                    # 表の前に少しスペースを開ける
                    current_row += 1 
                    for table_info in page_data['tables']:
                        table_data = table_info['data']
                        table_idx = table_info['table_index']
                        
                        # 表のタイトルを追加（複数の表がある場合）
                        if len(page_data['tables']) > 1:
                            table_title = f"表 {table_idx + 1}"
                            new_sheet.cell(row=current_row, column=1, value=table_title)
                            current_row += 1
                        
                        # 表のデータをExcelセルに配置
                        for row_idx, row_data in enumerate(table_data):
                            for col_idx, cell_data in enumerate(row_data):
                                if cell_data:  # 空でないセルのみ配置
                                    # セルの文字数制限を考慮
                                    cell_text = str(cell_data)
                                    if len(cell_text) > 32767:
                                        cell_text = cell_text[:32767]
                                    
                                    new_sheet.cell(
                                        row=current_row + row_idx, 
                                        column=col_idx + 1, 
                                        value=cell_text
                                    )
                        
                        # 表の後に空行を追加
                        current_row += len(table_data) + 2
                
                # 表以外のテキストがある場合は追加情報として配置
                # 抽出済みのテキストはここには含まれないように調整済み
                if page_data['text'] and page_data['text'].strip():
                    # テキストセクションのヘッダー（表がある場合は追加）
                    # 「その他のテキスト」というヘッダーは、実際に残りのテキストがある場合にのみ表示
                    new_sheet.cell(row=current_row, column=1, value="その他のテキスト:")
                    current_row += 1
                    
                    # テキストを行ごとに分割して配置
                    text_lines = page_data['text'].split('\n')
                    for line in text_lines:
                        if line.strip():  # 空行はスキップ
                            # セルの文字数制限を考慮
                            if len(line) > 32767:
                                line = line[:32767]
                            
                            new_sheet.cell(row=current_row, column=1, value=line.strip())
                            current_row += 1

        # 6. 処理済みファイルをメモリに保存
        output_excel_stream = io.BytesIO()
        
        # xlsmファイルとして保存（VBAマクロも保持）
        template_workbook.save(output_excel_stream)
        output_excel_stream.seek(0) # ストリームの先頭に戻す

        # 7. 処理済みExcelファイルをダウンロード用に返す
        download_filename = f'processed_template_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsm'
        
        return send_file(
            output_excel_stream,
            mimetype='application/vnd.ms-excel.sheet.macroEnabled.12',  # xlsmのMIMEタイプ
            as_attachment=True,
            download_name=download_filename
        )

    except Exception as e:
        # エラーログ出力
        print(f"An error occurred: {e}")
        traceback.print_exc() # 詳細なトレースバックを出力

        return render_template('error.html', message=f"ファイル処理中に予期せぬエラーが発生しました: {e}"), 500

if __name__ == '__main__':
    app.run(debug=True)
